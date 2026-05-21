# code = utf-8
from openpyxl import load_workbook
import numpy as np

import matplotlib.pyplot as plt

# 加载工作簿 表格名称
workbook = load_workbook(filename='Microsoft Excel 工作表 (2).xlsx')

# 选择工作表
sheet = workbook.active
# arr = np.array(sheet)
# print(arr.shape)
row = sheet.max_row
print("row:%d" % sheet.max_row)  # test.xlsx 表格1的行数
print("col:%d" % sheet.max_column)  # test.xlsx 表格1的列数

col1 = []
col2 = []

for i in range(row):
    # 第一列
    col1.append(sheet.cell(i + 1, 1).value)
    # 第二列
    col2.append(sheet.cell(i + 1, 2).value)

# 打印索引和值
for line_idx, line in enumerate(col1):
    print(line)

plt.plot(col1)
plt.plot(col2)

# plt.legend()

# 读取每一行打印
# for row in sheet.iter_rows(values_only=True):
#    print(row)

# 范围显示
# plt.xlim(0, 200)
plt.show()